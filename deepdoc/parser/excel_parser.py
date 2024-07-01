# 导入load_workbook函数，用于加载Excel工作簿
from openpyxl import load_workbook
# 导入sys模块，用于系统相关的功能
import sys
# 导入BytesIO类，用于创建内存中的二进制文件对象
from io import BytesIO

# 导入find_codec函数，用于查找并返回合适的编码方式
from rag.nlp import find_codec


class RAGFlowExcelParser:
    def html(self, fnm, chunk_rows=256):
        """
        将Excel工作簿转换为HTML表格的列表。

        每个Excel工作表被转换为一个HTML表格，如果工作表中的行数超过chunk_rows参数指定的数量，
        则会被拆分为多个HTML表格块。

        参数:
        - fnm: Excel文件名或包含Excel数据的字节流。
        - chunk_rows: 每个HTML表格块包含的最大行数。

        返回:
        - 一个包含HTML表格字符串的列表，每个Excel工作表对应一个或多个表格。
        """
        # 根据输入参数的类型，加载Excel工作簿
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))

        # 存储所有工作表的HTML表格块
        tb_chunks = []
        # 遍历工作簿中的每个工作表
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            # 如果工作表为空，则跳过
            if not rows: continue

            # 构建工作表的表头行
            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            # 将工作表拆分为多个表格块
            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                # 开始构建一个表格块，包括表名和表头
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                # 构建表格的主体部分
                for r in list(rows[1 + chunk_i * chunk_rows:1 + (chunk_i + 1) * chunk_rows]):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        # 对于空单元格，使用空td标签
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                # 表格块构建完成，添加到结果列表中
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
        """
        将实例作为函数调用，用于处理Excel文件中的数据。

        参数:
        fnm: 字符串类型或BytesIO对象，表示Excel文件的路径或内存中的Excel数据。

        返回:
        list: 包含处理后数据的字符串列表，每个字符串代表一个单元格的数据。
        """
        # 根据参数类型加载工作簿，支持从文件路径和BytesIO对象加载
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))

        # 用于存储处理后的数据
        res = []

        # 遍历工作簿中的所有工作表
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)

            # 如果工作表中没有数据行，则跳过
            if not rows:
                continue

            # 获取第一行的数据作为标题
            ti = list(rows[0])

            # 遍历除第一行外的所有数据行
            for r in list(rows[1:]):
                l = []

                # 遍历当前行的每个单元格
                for i, c in enumerate(r):
                    # 如果单元格值为空，则跳过
                    if not c.value:
                        continue

                    # 构建单元格的字符串表示，如果当前单元格有标题，则加上标题
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    l.append(t)

                # 将当前行的数据用"; "连接成一个字符串，并根据工作表名称添加后缀
                l = "; ".join(l)
                if sheetname.lower().find("sheet") < 0:
                    l += " ——" + sheetname
                res.append(l)

        return res

    @staticmethod
    def row_number(fnm, binary):
        """
        计算给定二进制数据表示的文件中的总行数。

        该方法能够处理Excel、CSV和TXT格式的文件。对于Excel文件，它计算所有工作表的行数总和；
        对于CSV和TXT文件，它计算文本中的行数。

        参数:
        fnm (str): 文件名，用于判断文件类型。
        binary (bytes): 文件的二进制数据。

        返回:
        int: 文件中的总行数。
        """
        # 判断文件是否为Excel格式
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            # 使用BytesIO将二进制数据模拟为文件，加载Excel工作簿
            wb = load_workbook(BytesIO(binary))
            total = 0
            # 遍历所有工作表，累加行数
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                # 计算当前工作表的行数并累加到总数
                total += len(list(ws.rows))
            return total

        # 判断文件是否为CSV或TXT格式
        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            # 确定文本的编码方式
            encoding = find_codec(binary)
            # 解码二进制数据，忽略错误
            txt = binary.decode(encoding, errors="ignore")
            # 计算并返回文本的行数
            return len(txt.split("\n"))


# 当模块作为主程序运行时，执行以下代码
if __name__ == "__main__":
    # 初始化RAGFlowExcelParser对象，用于处理Excel文件
    psr = RAGFlowExcelParser()
    # 调用对象的解析方法，传入命令行参数中的第一个参数作为文件路径
    psr(sys.argv[1])

